import streamlit as st
import pdfplumber
import re
import io
import pandas as pd
import plotly.graph_objects as go
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Page config ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Journal des Appels — Hympyr",
    page_icon="📞",
    layout="wide",
)

st.markdown("""
<style>
    .stApp { font-family: 'Segoe UI', sans-serif; background: #f4f6fb; }
    h1, h2, h3 { color: #1a237e; }
    .metric-card {
        background: white; border-radius: 14px; padding: 22px 12px;
        box-shadow: 0 2px 10px rgba(0,0,0,0.07); text-align: center;
    }
    .metric-value { font-size: 2.1rem; font-weight: 800; }
    .metric-label { font-size: 0.82rem; color: #777; margin-top: 5px;
        text-transform: uppercase; letter-spacing: .04em; }
    .section-title {
        font-size: 1.15rem; font-weight: 700; color: #1a237e;
        border-left: 4px solid #e65100; padding-left: 10px;
        margin: 28px 0 14px 0;
    }
    .alert-badge {
        background: #fff3e0; border: 1.5px solid #e65100; border-radius: 8px;
        padding: 12px 16px; font-size: 0.9rem; color: #bf360c; line-height: 1.7;
    }
    .client-urgent {
        background: #fce4ec; border-left: 4px solid #c62828;
        padding: 10px 14px; border-radius: 6px;
        font-size: 0.88rem; line-height: 1.5;
    }
</style>
""", unsafe_allow_html=True)

NAVY   = "#1a237e"
ORANGE = "#e65100"
GREEN  = "#2e7d32"
RED    = "#c62828"

# ─── Helpers ─────────────────────────────────────────────────────────────────
def duration_to_seconds(d: str) -> int:
    if not d:
        return 0
    parts = d.split(":")
    try:
        return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
    except Exception:
        return 0

def fmt_seconds(s: int) -> str:
    m, sec = divmod(int(s), 60)
    return f"{m}m {sec:02d}s"

# ─── PDF Parsing ──────────────────────────────────────────────────────────────
def parse_pdf(file_bytes: bytes) -> tuple[list[dict], dict]:
    records = []
    metadata = {}
    re_answered   = re.compile(r'^X\s+(\d{2}:\d{2})\s+(\d{2}:\d{2})\s+(.+?)\s+(\d[\d\s]{9,14})\s+(\d+:\d{2}:\d{2})\s*$')
    re_unanswered = re.compile(r'^(\d{2}:\d{2})\s+(.+?)\s+(\d[\d\s]{9,14})\s*$')
    re_date       = re.compile(r'du:\s*(\d{2}/\d{2}/\d{4})')
    re_society    = re.compile(r'Société:\s*(\S+)')

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for raw_line in text.splitlines():
                line = raw_line.strip()
                if not line:
                    continue
                m = re_date.search(line);    
                if m: metadata.setdefault("date", m.group(1))
                m = re_society.search(line); 
                if m: metadata.setdefault("societe", m.group(1))
                m = re.search(r"Nombre d'Appels Répondus\s+(\d+)", line)
                if m: metadata["repondus"] = int(m.group(1))
                m = re.search(r"Nombre d'Appels Sans Réponse\s+(\d+)", line)
                if m: metadata["sans_reponse"] = int(m.group(1))
                m = re.search(r"Nombre d'Appels Total\s+(\d+)", line)
                if m: metadata["total"] = int(m.group(1))

                m = re_answered.match(line)
                if m:
                    debut, fin, adresse, telephone, duree = m.groups()
                    telephone = telephone.strip()
                    absent = "<Absent" in adresse
                    clean_addr = re.sub(r'\s*<Absent[^>]*>', '', adresse).strip()
                    records.append({
                        "Statut": "Répondu",
                        "Heure début": debut,
                        "Heure fin": fin,
                        "Heure (int)": int(debut[:2]),
                        "Adresse / Client": clean_addr,
                        "Téléphone": telephone,
                        "Durée": duree,
                        "Durée (s)": duration_to_seconds(duree),
                        "Absent répertoire": "Oui" if absent else "Non",
                    })
                    continue

                m = re_unanswered.match(line)
                if m:
                    debut, adresse, telephone = m.groups()
                    telephone = telephone.strip()
                    absent = "<Absent" in adresse
                    clean_addr = re.sub(r'\s*<Absent[^>]*>', '', adresse).strip()
                    records.append({
                        "Statut": "Sans réponse",
                        "Heure début": debut,
                        "Heure fin": "",
                        "Heure (int)": int(debut[:2]),
                        "Adresse / Client": clean_addr,
                        "Téléphone": telephone,
                        "Durée": "",
                        "Durée (s)": 0,
                        "Absent répertoire": "Oui" if absent else "Non",
                    })
    return records, metadata

# ─── Analytics ───────────────────────────────────────────────────────────────
def build_hourly_df(df: pd.DataFrame) -> pd.DataFrame:
    all_hours = list(range(df["Heure (int)"].min(), df["Heure (int)"].max() + 1))
    rep = df[df["Statut"] == "Répondu"].groupby("Heure (int)").size()
    sr  = df[df["Statut"] == "Sans réponse"].groupby("Heure (int)").size()
    result = pd.DataFrame({"Heure": all_hours})
    result["Répondus"]     = result["Heure"].map(rep).fillna(0).astype(int)
    result["Sans réponse"] = result["Heure"].map(sr).fillna(0).astype(int)
    result["Total"]        = result["Répondus"] + result["Sans réponse"]
    result["Taux réponse"] = (result["Répondus"] / result["Total"].replace(0, 1) * 100).round(1)
    result["Label"]        = result["Heure"].apply(lambda h: f"{h:02d}h")
    return result

def build_client_df(df: pd.DataFrame) -> pd.DataFrame:
    grp = df.groupby("Téléphone").agg(
        Nom=("Adresse / Client", "first"),
        Appels=("Statut", "count"),
        Répondus=("Statut", lambda x: (x == "Répondu").sum()),
        Durée_totale_s=("Durée (s)", "sum"),
        Première_heure=("Heure début", "min"),
        Dernière_heure=("Heure début", "max"),
    ).reset_index()
    grp["Sans réponse"]   = grp["Appels"] - grp["Répondus"]
    grp["Durée totale"]   = grp["Durée_totale_s"].apply(fmt_seconds)
    grp["Taux réponse %"] = (grp["Répondus"] / grp["Appels"] * 100).round(0).astype(int)
    return grp.sort_values("Appels", ascending=False).reset_index(drop=True)

# ─── Charts ───────────────────────────────────────────────────────────────────
BASE_LAYOUT = dict(
    font_family="Segoe UI",
    paper_bgcolor="white",
    plot_bgcolor="white",
    margin=dict(l=20, r=20, t=44, b=20),
)

def chart_hourly_bars(hourly: pd.DataFrame) -> go.Figure:
    fig = go.Figure()
    fig.add_trace(go.Bar(
        name="Répondus", x=hourly["Label"], y=hourly["Répondus"],
        marker_color=GREEN,
        text=hourly["Répondus"], textposition="inside",
        textfont=dict(color="white", size=10),
    ))
    fig.add_trace(go.Bar(
        name="Sans réponse", x=hourly["Label"], y=hourly["Sans réponse"],
        marker_color=RED,
        text=hourly["Sans réponse"], textposition="inside",
        textfont=dict(color="white", size=10),
    ))
    fig.add_trace(go.Scatter(
        name="Taux de réponse (%)",
        x=hourly["Label"], y=hourly["Taux réponse"],
        mode="lines+markers", yaxis="y2",
        line=dict(color=ORANGE, width=2.5, dash="dot"),
        marker=dict(size=7, color=ORANGE),
    ))
    fig.update_layout(
        **BASE_LAYOUT,
        title=dict(text="Charge horaire & taux de réponse", font=dict(size=14, color=NAVY)),
        barmode="stack",
        legend=dict(orientation="h", y=-0.2),
        yaxis=dict(title="Nombre d'appels", gridcolor="#eeeeee"),
        yaxis2=dict(
            title="Taux (%)", overlaying="y", side="right",
            range=[0, 115], ticksuffix="%", showgrid=False,
        ),
        height=390,
    )
    return fig

def chart_heatmap_intensity(hourly: pd.DataFrame) -> go.Figure:
    z = [hourly["Total"].tolist()]
    x = hourly["Label"].tolist()
    fig = go.Figure(go.Heatmap(
        z=z, x=x, y=[""],
        colorscale=[[0, "#e8f5e9"], [0.35, "#66bb6a"], [0.65, "#e65100"], [1, "#b71c1c"]],
        showscale=True,
        colorbar=dict(title="Appels", thickness=12, len=0.8),
        text=[[str(v) for v in hourly["Total"].tolist()]],
        texttemplate="%{text}",
        textfont=dict(size=12, color="white"),
    ))
    layout = dict(**BASE_LAYOUT)
    layout["margin"] = dict(l=20, r=20, t=44, b=10)
    fig.update_layout(
        **layout,
        title=dict(text="Heatmap d'intensité horaire", font=dict(size=14, color=NAVY)),
        height=160,
    )
    return fig

def chart_top_clients(client_df: pd.DataFrame, n: int = 15) -> go.Figure:
    top = client_df.head(n).copy().sort_values("Appels")
    top["Nom court"] = top["Nom"].apply(lambda x: (x[:32] + "…") if len(x) > 32 else x)
    fig = go.Figure()
    fig.add_trace(go.Bar(
        name="Répondus", y=top["Nom court"], x=top["Répondus"],
        orientation="h", marker_color=GREEN,
        text=top["Répondus"], textposition="inside",
        textfont=dict(color="white", size=10),
    ))
    fig.add_trace(go.Bar(
        name="Sans réponse", y=top["Nom court"], x=top["Sans réponse"],
        orientation="h", marker_color=RED,
        text=top["Sans réponse"].apply(lambda v: str(v) if v > 0 else ""),
        textposition="inside", textfont=dict(color="white", size=10),
    ))
    fig.update_layout(
        **BASE_LAYOUT,
        title=dict(text=f"Top {n} appelants", font=dict(size=14, color=NAVY)),
        barmode="stack",
        height=max(360, n * 32),
        legend=dict(orientation="h", y=-0.1),
        xaxis=dict(title="Nombre d'appels", gridcolor="#eeeeee"),
        yaxis=dict(tickfont=dict(size=10)),
    )
    return fig

def chart_bubble_multi(client_df: pd.DataFrame) -> go.Figure | None:
    multi = client_df[client_df["Appels"] > 1].copy()
    if multi.empty:
        return None
    multi["Nom court"]    = multi["Nom"].apply(lambda x: (x[:26] + "…") if len(x) > 26 else x)
    multi["bubble_size"]  = (multi["Durée_totale_s"] / 25).clip(lower=8, upper=55)
    fig = go.Figure(go.Scatter(
        x=multi["Appels"],
        y=multi["Taux réponse %"],
        mode="markers+text",
        text=multi["Nom court"],
        textposition="top center",
        textfont=dict(size=9),
        marker=dict(
            size=multi["bubble_size"],
            color=multi["Taux réponse %"],
            colorscale=[[0, RED], [0.5, ORANGE], [1, GREEN]],
            showscale=True,
            colorbar=dict(title="Taux<br>réponse %", thickness=12, len=0.6),
            line=dict(color="white", width=1.5),
        ),
        hovertemplate=(
            "<b>%{text}</b><br>Appels : %{x}<br>Taux réponse : %{y}%<extra></extra>"
        ),
    ))
    fig.add_hline(y=50, line_dash="dot", line_color=ORANGE,
                  annotation_text="seuil 50 %", annotation_font_size=10)
    fig.update_layout(
        **BASE_LAYOUT,
        title=dict(text="Clients multi-appels — Intensité vs qualité de traitement",
                   font=dict(size=14, color=NAVY)),
        xaxis=dict(title="Appels dans la journée", gridcolor="#eeeeee", dtick=1),
        yaxis=dict(title="Taux de réponse (%)", range=[-5, 115], gridcolor="#eeeeee"),
        height=430,
    )
    return fig

# ─── Excel Builder ────────────────────────────────────────────────────────────
def build_excel(records: list[dict], metadata: dict, date_str: str) -> bytes:
    wb = Workbook()
    NH = "1A237E"; OH = "E65100"; GH = "2E7D32"; RH = "C62828"
    WH = "FFFFFF"; GB = "F5F5F5"; GL = "BDBDBD"; LB = "EEF2FF"
    thin  = Side(border_style="thin",   color=GL)
    thick = Side(border_style="medium", color=NH)
    bd = Border(left=thin, right=thin, top=thin,  bottom=thin)
    bt = Border(left=thin, right=thin, top=thick, bottom=thin)
    bb = Border(left=thin, right=thin, top=thin,  bottom=thick)
    def hf(sz=10, c=WH): return Font(name="Arial", bold=True, size=sz, color=c)
    def cf(sz=9, b=False, c="212121"): return Font(name="Arial", bold=b, size=sz, color=c)
    def fl(h): return PatternFill("solid", fgColor=h)

    # ── Sheet 1 – Journal ─────────────────────────────────────────────────────
    ws = wb.active; ws.title = "Journal des appels"
    ws.sheet_view.showGridLines = False
    c = ws.cell(row=1, column=1)
    c.value = f"HYMPYR — Journal des Appels  ·  {date_str}"
    c.font = hf(14); c.fill = fl(NH)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 36

    rep_n  = metadata.get("repondus", sum(1 for r in records if r["Statut"]=="Répondu"))
    tot_n  = metadata.get("total", len(records))
    kpis = [
        (1, 2, f"Total : {tot_n}", NH),
        (3, 4, f"Répondus : {rep_n}", GH),
        (5, 6, f"Sans rép. : {tot_n-rep_n}", RH),
        (7, 7, f"Taux : {rep_n/max(tot_n,1)*100:.1f}%", OH),
    ]
    for col_start, col_end, txt, fc in kpis:
        cell = ws.cell(row=2, column=col_start)
        cell.value = txt; cell.font = hf()
        cell.fill = fl(fc)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_end > col_start:
            ws.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_end)
    ws.row_dimensions[2].height = 26; ws.row_dimensions[3].height = 4

    headers    = ["Statut","Heure début","Heure fin","Adresse / Client","Téléphone","Durée","Absent répertoire"]
    col_widths = [14, 12, 10, 44, 18, 10, 18]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(4, ci, h); c.font = hf(); c.fill = fl(NH)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = bt
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[4].height = 22; ws.freeze_panes = "A5"

    for ri, rec in enumerate(records, 5):
        answered = rec["Statut"] == "Répondu"
        bg = fl(WH) if ri % 2 == 0 else fl(GB)
        for ci, key in enumerate(headers, 1):
            val = rec.get(key, "")
            cell = ws.cell(ri, ci, val)
            cell.fill = bg; cell.border = bd; cell.font = cf()
            cell.alignment = Alignment(vertical="center")
            if ci == 1:
                cell.font = cf(b=True, c=GH if answered else RH)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if ci == 5:
                cell.font = Font(name="Courier New", size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if ci in (2, 3, 6):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if ci == 7 and val == "Oui":
                cell.font = cf(c=RH)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[ri].height = 17

    total_row = len(records) + 5
    ct = ws.cell(row=total_row, column=1)
    ct.value = f"TOTAL : {len(records)} appels"
    # merge after setting value
    ws.merge_cells(f"A{total_row}:G{total_row}")
    ct.font = hf(); ct.fill = fl(NH)
    ct.alignment = Alignment(horizontal="center", vertical="center")
    ct.border = bb; ws.row_dimensions[total_row].height = 20
    ws.auto_filter.ref = f"A4:G{total_row-1}"

    # ── Sheet 2 – Synthèse ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Synthèse"); ws2.sheet_view.showGridLines = False
    c2t = ws2.cell(row=1, column=1)
    c2t.value = "Synthèse — Journal des Appels"
    c2t.font = hf(13); c2t.fill = fl(NH)
    c2t.alignment = Alignment(horizontal="center", vertical="center")
    ws2.merge_cells("A1:D1")
    ws2.row_dimensions[1].height = 32
    for col, w in [("A",2),("B",30),("C",20),("D",2)]:
        ws2.column_dimensions[col].width = w

    durations = [r["Durée (s)"] for r in records if r["Durée (s)"] > 0]
    avg_sec   = int(sum(durations)/len(durations)) if durations else 0
    kpi_rows  = [
        ("Date du journal",     date_str),
        ("Total appels",        tot_n),
        ("Appels répondus",     rep_n),
        ("Appels sans réponse", tot_n - rep_n),
        ("Taux de réponse",     rep_n / max(tot_n, 1)),
        ("Durée moy. répondus", fmt_seconds(avg_sec)),
    ]
    for i, (lbl, val) in enumerate(kpi_rows, 2):
        lc = ws2.cell(i+1, 2, lbl)
        lc.font = cf(10, True); lc.fill = fl(LB)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1); lc.border = bd
        vc = ws2.cell(i+1, 3, val)
        vc.font = cf(10); vc.fill = fl(WH)
        vc.alignment = Alignment(horizontal="center", vertical="center"); vc.border = bd
        ws2.row_dimensions[i+1].height = 24
        if lbl == "Taux de réponse":
            vc.number_format = "0.0%"; vc.font = cf(10, True, GH)

    # Répartition horaire
    ho, hk = Counter(), Counter()
    for r in records:
        h = r["Heure début"][:2] if r["Heure début"] else None
        if h: (ho if r["Statut"]=="Répondu" else hk)[h] += 1
    all_h = sorted(set(list(ho.keys()) + list(hk.keys())))
    sr = 11
    csr = ws2.cell(row=sr, column=2)
    csr.value = "Répartition horaire"
    csr.font = hf(); csr.fill = fl(OH)
    csr.alignment = Alignment(horizontal="center"); csr.border = bd
    ws2.merge_cells(f"B{sr}:D{sr}")
    for j, h in enumerate(["Heure","Répondus","Sans réponse"], 2):
        c = ws2.cell(sr+1, j, h); c.font = hf(); c.fill = fl(NH)
        c.alignment = Alignment(horizontal="center"); c.border = bd
    for k, h in enumerate(all_h, sr+2):
        ws2.cell(k,2,f"{h}h").font = cf(b=True)
        ws2.cell(k,2).fill = fl(GB); ws2.cell(k,2).alignment = Alignment(horizontal="center"); ws2.cell(k,2).border = bd
        for col, val in [(3, ho.get(h,0)), (4, hk.get(h,0))]:
            ws2.cell(k,col,val).fill = fl(WH)
            ws2.cell(k,col).alignment = Alignment(horizontal="center"); ws2.cell(k,col).border = bd
        ws2.row_dimensions[k].height = 18

    # ── Sheet 3 – Multi-appels ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Clients multi-appels"); ws3.sheet_view.showGridLines = False
    c3t = ws3.cell(row=1, column=1)
    c3t.value = "Clients ayant appelé plusieurs fois dans la journée"
    c3t.font = hf(12); c3t.fill = fl(OH)
    c3t.alignment = Alignment(horizontal="center", vertical="center"); ws3.row_dimensions[1].height = 30
    ws3.merge_cells("A1:F1")
    df_tmp   = pd.DataFrame(records)
    cli_grp  = df_tmp.groupby("Téléphone").agg(
        Nom=("Adresse / Client","first"),
        Appels=("Statut","count"),
        Rep=("Statut", lambda x: (x=="Répondu").sum()),
        Dur=("Durée (s)","sum"),
    ).reset_index()
    cli_grp = cli_grp[cli_grp["Appels"] > 1].sort_values("Appels", ascending=False)
    hdrs3 = ["Client","Téléphone","Nb appels","Répondus","Sans réponse","Durée totale"]
    wds3  = [38, 18, 12, 12, 14, 14]
    for ci, (h, w) in enumerate(zip(hdrs3, wds3), 1):
        c = ws3.cell(2, ci, h); c.font = hf(); c.fill = fl(NH)
        c.alignment = Alignment(horizontal="center"); c.border = bd
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.row_dimensions[2].height = 22
    for ri, (_, row) in enumerate(cli_grp.iterrows(), 3):
        bg = fl(WH) if ri%2==0 else fl(GB)
        vals = [row["Nom"], row["Téléphone"], row["Appels"], row["Rep"],
                row["Appels"]-row["Rep"], fmt_seconds(int(row["Dur"]))]
        for ci, val in enumerate(vals, 1):
            cell = ws3.cell(ri, ci, val); cell.fill = bg; cell.border = bd; cell.font = cf()
            cell.alignment = Alignment(vertical="center", horizontal="center" if ci > 1 else "left")
        ws3.row_dimensions[ri].height = 17

    # ── Sheet 4 – Numéros inconnus ────────────────────────────────────────────
    ws4 = wb.create_sheet("Numéros inconnus"); ws4.sheet_view.showGridLines = False
    c4t = ws4.cell(row=1, column=1)
    c4t.value = "Numéros absents du répertoire"
    c4t.font = hf(12); c4t.fill = fl(RH)
    c4t.alignment = Alignment(horizontal="center", vertical="center"); ws4.row_dimensions[1].height = 30
    ws4.merge_cells("A1:C1")
    unknowns = [r for r in records if r["Absent répertoire"] == "Oui"]
    for j, h in enumerate(["Statut","Heure début","Téléphone"], 1):
        c = ws4.cell(2,j,h); c.font=hf(); c.fill=fl(NH)
        c.alignment=Alignment(horizontal="center"); c.border=bd
    for col, w in [("A",14),("B",13),("C",20)]:
        ws4.column_dimensions[col].width = w
    for i, r in enumerate(unknowns, 3):
        answered = r["Statut"]=="Répondu"
        bg = fl(WH) if i%2==0 else fl(GB)
        ws4.cell(i,1,r["Statut"]).font = Font(name="Arial",size=9,bold=True,color=GH if answered else RH)
        ws4.cell(i,1).fill=bg; ws4.cell(i,1).alignment=Alignment(horizontal="center"); ws4.cell(i,1).border=bd
        ws4.cell(i,2,r["Heure début"]).fill=bg; ws4.cell(i,2).alignment=Alignment(horizontal="center"); ws4.cell(i,2).border=bd
        ws4.cell(i,3,r["Téléphone"]).font=Font(name="Courier New",size=9)
        ws4.cell(i,3).fill=bg; ws4.cell(i,3).alignment=Alignment(horizontal="center"); ws4.cell(i,3).border=bd
        ws4.row_dimensions[i].height=18

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

# ═════════════════════════════════════════════════════════════════════════════
# UI
# ═════════════════════════════════════════════════════════════════════════════
st.title("📞 Journal des Appels — Hympyr Energies")
st.markdown("Importez un **PDF Logimatique** pour obtenir analyses visuelles + export Excel professionnel.")
st.divider()

uploaded = st.file_uploader(
    "Déposer le fichier PDF ici", type=["pdf"],
    help="Journal des appels exporté depuis Logimatique (format HY / HYMPYR)",
)

if not uploaded:
    st.info("👆 Importez un fichier PDF pour commencer.")
    st.markdown("""
    **Ce que vous obtiendrez :**
    - KPIs synthétiques (total, répondus, taux de réponse, durée moyenne)
    - Barres empilées par heure + courbe du taux de réponse
    - Heatmap thermique d'intensité horaire
    - Top appelants + détection des clients urgents (≥ 3 appels)
    - Bubble chart multi-appels : intensité × qualité de traitement
    - Export Excel 4 onglets : Journal · Synthèse · Multi-appels · Inconnus
    """)
    st.stop()

with st.spinner("Analyse du PDF…"):
    records, metadata = parse_pdf(uploaded.read())

if not records:
    st.error("Aucun appel détecté. Vérifiez que le fichier est un Journal des Appels Logimatique.")
    st.stop()

df        = pd.DataFrame(records)
total     = metadata.get("total", len(records))
repondus  = metadata.get("repondus", int((df["Statut"]=="Répondu").sum()))
sans_rep  = metadata.get("sans_reponse", total - repondus)
taux      = repondus / max(total, 1) * 100
date_str  = metadata.get("date", "—")
durations = df[df["Durée (s)"]>0]["Durée (s)"]
avg_dur   = fmt_seconds(int(durations.mean())) if not durations.empty else "—"

hourly_df = build_hourly_df(df)
client_df = build_client_df(df)
multi_df  = client_df[client_df["Appels"] > 1].copy()

# ── KPI Cards ─────────────────────────────────────────────────────────────────
c1, c2, c3, c4, c5 = st.columns(5)
for col, label, val, color in [
    (c1, "Total appels",     total,             NAVY),
    (c2, "Répondus",        repondus,           GREEN),
    (c3, "Sans réponse",    sans_rep,           RED),
    (c4, "Taux de réponse", f"{taux:.1f} %",    ORANGE),
    (c5, "Durée moy.",      avg_dur,            "#00695c"),
]:
    col.markdown(f"""
    <div class="metric-card">
        <div class="metric-value" style="color:{color}">{val}</div>
        <div class="metric-label">{label}</div>
    </div>""", unsafe_allow_html=True)

st.markdown(f"<br>**Date :** {date_str} &nbsp;|&nbsp; **Société :** {metadata.get('societe','—')}",
            unsafe_allow_html=True)
st.divider()

# ═══════════════════════════════════════════════════════════════
# SECTION 1 — ANALYSE TEMPORELLE
# ═══════════════════════════════════════════════════════════════
st.markdown('<div class="section-title">⏱ Analyse temporelle</div>', unsafe_allow_html=True)

col_a, col_b = st.columns([3, 1])
with col_a:
    st.plotly_chart(chart_hourly_bars(hourly_df), use_container_width=True)
with col_b:
    peak  = hourly_df.loc[hourly_df["Total"].idxmax()]
    worst = hourly_df.loc[hourly_df["Taux réponse"].idxmin()]
    st.markdown("**Insights clés**")
    st.markdown(f"""
    <div class="alert-badge">
    🔴 <b>Pic de charge :</b><br>{peak['Label']} — {int(peak['Total'])} appels<br><br>
    ⚠️ <b>Pire taux de réponse :</b><br>{worst['Label']} — {worst['Taux réponse']:.0f}%
    &nbsp;({int(worst['Sans réponse'])} non traités)
    </div>""", unsafe_allow_html=True)
    st.markdown("")
    st.dataframe(
        hourly_df[["Label","Répondus","Sans réponse","Taux réponse"]]
        .rename(columns={"Label":"Heure","Taux réponse":"Taux (%)"}),
        hide_index=True, use_container_width=True, height=220,
    )

st.plotly_chart(chart_heatmap_intensity(hourly_df), use_container_width=True)
st.divider()

# ═══════════════════════════════════════════════════════════════
# SECTION 2 — ANALYSE CLIENTS & MULTI-APPELS
# ═══════════════════════════════════════════════════════════════
st.markdown('<div class="section-title">👥 Analyse clients & multi-appels</div>', unsafe_allow_html=True)

# Alerte clients urgents ≥ 3 appels
urgents = multi_df[multi_df["Appels"] >= 3].sort_values("Appels", ascending=False)
if not urgents.empty:
    st.markdown(f"⚠️ **{len(urgents)} client(s) ont appelé ≥ 3 fois** — à recontacter en priorité :")
    cols_u = st.columns(min(len(urgents), 4))
    for i, (_, row) in enumerate(urgents.iterrows()):
        if i >= 4: break
        cols_u[i].markdown(f"""
        <div class="client-urgent">
        <b>{row['Nom'][:28]}</b><br>
        📞 {row['Téléphone']}<br>
        {row['Appels']} appels · {row['Taux réponse %']}% répondus
        </div>""", unsafe_allow_html=True)
    st.markdown("")

col_c, col_d = st.columns(2)
with col_c:
    n_top = st.slider("Nombre de clients à afficher", 5, 30, 15)
    st.plotly_chart(chart_top_clients(client_df, n_top), use_container_width=True)
with col_d:
    bubble = chart_bubble_multi(client_df)
    if bubble:
        st.plotly_chart(bubble, use_container_width=True)
    else:
        st.info("Aucun client multi-appels détecté.")

if not multi_df.empty:
    with st.expander(f"📋 Détail des {len(multi_df)} clients multi-appels"):
        st.dataframe(
            multi_df[["Nom","Téléphone","Appels","Répondus","Sans réponse","Durée totale","Taux réponse %","Première_heure","Dernière_heure"]]
            .rename(columns={"Première_heure":"1er appel","Dernière_heure":"Dernier appel"}),
            hide_index=True, use_container_width=True, height=300,
        )

st.divider()

# ═══════════════════════════════════════════════════════════════
# SECTION 3 — JOURNAL BRUT
# ═══════════════════════════════════════════════════════════════
st.markdown('<div class="section-title">📋 Journal des appels brut</div>', unsafe_allow_html=True)

display_df = df.drop(columns=["Heure (int)", "Durée (s)"])
tab1, tab2, tab3 = st.tabs(["Tous les appels", "✅ Répondus", "❌ Sans réponse"])
with tab1: st.dataframe(display_df, use_container_width=True, height=360)
with tab2: st.dataframe(display_df[display_df["Statut"]=="Répondu"], use_container_width=True, height=360)
with tab3: st.dataframe(display_df[display_df["Statut"]=="Sans réponse"], use_container_width=True, height=360)

st.divider()

with st.spinner("Génération Excel…"):
    xlsx_bytes = build_excel(records, metadata, date_str)

filename = f"Journal_Appels_HYMPYR_{date_str.replace('/','')}.xlsx"
st.download_button(
    label="⬇️  Télécharger le fichier Excel (4 onglets)",
    data=xlsx_bytes,
    file_name=filename,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
    type="primary",
)
st.caption(f"**{filename}** · 4 onglets : Journal · Synthèse · Multi-appels · Numéros inconnus")
